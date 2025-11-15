# ---------------------------- import ------------------------------- #
import html
import pandas as pd
import requests
import concurrent.futures
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox


# ---------------------------- __init__ ------------------------------- #
class DataRework:
    def __init__(self, excel_user_input):
        self.excel = None
        self.sheet = None
        self.excel_file_input = excel_user_input
        self.convert_excel()
        self.add_column()
        self.url = r"https://openplzapi.org/de/"
        self.appending_data = {}

# ---------------------------- api ------------------------------- #

    def fetch_data(self, plz_state):
        plz, state = plz_state
        url = f"{self.url}Localities?postalCode={plz}"
        response = requests.get(url)
        response.raise_for_status()
        response_data = response.json()

        for item in response_data:
            municipality_name = item["municipality"]["name"]
            if re.split(r",|-|\s", municipality_name)[0] == re.split(r",|-|\s", state)[0]:
                federal_state = item.get("federalState", {}).get("name", "")
                district = item.get("district", {}).get("name", "")
                return federal_state, district, municipality_name
        return None

# ---------------------------- main process ------------------------------- #

    def check_plz_parallel(self, callback=None):
        updated_sheets = {}

        for sheet_name, sheet_df in self.excel.items():

            # --- 1. Cleaning ---
            sheet_df.replace([" NULL", "NULL", "0NULL"], "NULLDATA", inplace=True)

            # --- 2. Mask: PLZ fehlt, Stadt vorhanden → Stadt überall eintragen ---
            mask = (sheet_df["Postleitzahl"] == "NULLDATA") & (sheet_df["Stadt"] != "NULLDATA")

            sheet_df.loc[mask, "federalState"] = sheet_df.loc[mask, "Stadt"]
            sheet_df.loc[mask, "district"] = sheet_df.loc[mask, "Stadt"]
            sheet_df.loc[mask, "municipality"] = sheet_df.loc[mask, "Stadt"]

            # --- 3. Progress start ---
            if callback:
                callback(0, len(sheet_df), sheet_name)

            plz_state_list = [
                (row.Postleitzahl, html.unescape(str(row.Stadt)) if pd.notnull(row.Stadt) else "")
                for _, row in sheet_df.iterrows()
            ]

            # Ergebniscontainer
            federal_list = []
            district_list = []
            municipality_list = []

            total = len(plz_state_list)
            count = 0

            with concurrent.futures.ThreadPoolExecutor(max_workers=25) as executor:
                futures = {
                    executor.submit(self.fetch_data, plz_state): idx
                    for idx, plz_state in enumerate(plz_state_list)
                }

                results = [None] * len(plz_state_list)

                for future in concurrent.futures.as_completed(futures):
                    idx = futures[future]
                    results[idx] = future.result()
                    count += 1

                    if callback:
                        callback(count, total, sheet_name)

            # --- 4. Werte einbauen ---
            for i, result in enumerate(results):

                # Wenn MASK zutrifft → API ignorieren
                if mask.iloc[i]:
                    federal_list.append(sheet_df.iloc[i]["federalState"])
                    district_list.append(sheet_df.iloc[i]["district"])
                    municipality_list.append(sheet_df.iloc[i]["municipality"])
                    continue

                # Normale API-Daten:
                if result is None:
                    federal_state, district, municipality = "", "", ""
                else:
                    federal_state, district, municipality = [
                        x.split(",")[0] if x else "" for x in result
                    ]

                federal_list.append(federal_state)
                district_list.append(district)
                municipality_list.append(municipality)

            sheet_df["federalState"] = federal_list
            sheet_df["district"] = district_list
            sheet_df["municipality"] = municipality_list

            updated_sheets[sheet_name] = sheet_df

        self.excel = updated_sheets

    # ---------------------------- prepare data ------------------------------- #

    def convert_excel(self):
        """Try to load excel data"""
        try:
            self.excel = pd.read_excel(self.excel_file_input, sheet_name=None, engine="openpyxl", keep_default_na=False)
        except ValueError as convert_error:
            print(f"Error - {convert_error}")

    def add_column(self):
        """Eine neue Spalte hinzufügen."""
        column_list = ["federalState","district","municipality"]
        for sheet_name, sheet_df in self.excel.items():
            for column in column_list:
                if column not in sheet_df.columns:
                    new_column = pd.Series([None]*len(sheet_df), name=column)
                    sheet_df = pd.concat([sheet_df, new_column], axis=1)
            sheet_df = self.plz_test(sheet_df)
            self.excel[sheet_name] = sheet_df

    def plz_test(self, df):
        plz_fixed_list = [
            (f"0{plz}" if len(str(plz)) == 4 else plz)
        for plz in df["Postleitzahl"]
        ]
        df['Postleitzahl'] = plz_fixed_list
        return df

    def save_excel(self, output_file):
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, sheet_df, in self.excel.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
class A:
    def __init__(self, excel_user_input):
        self.excel = None
        self.sheet = None
        self.excel_file_input = excel_user_input
        try:
            self.excel = pd.read_excel(self.excel_file_input, sheet_name=None, engine="openpyxl", keep_default_na=False)
        except ValueError as convert_error:
            print(f"Error - {convert_error}")
        self.analysis()

    def analysis(self):

 # ---------------------------- Gesamt-Zähler initialisieren ------------------------------- #
        total = {
            "foreign": 0,
            "polen": 0,
            "berlin": 0,
            "ffo": 0,
            "schoeneiche": 0,
            "erkner": 0,
            "woltersdorf": 0,
            "gruenheide": 0,
            "fuewa": 0,
            "spreenhagen": 0,
            "bad_saarow": 0,
            "storkow": 0,
            "rietz_neuendorf": 0,
            "tauche": 0,
            "odervorland": 0,
            "NULLDATA": 0,
            "no_category": 0,
            "sonstiges Deutschland": 0
        }
        results = []
        for sheet_name, sheet_df in self.excel.items():
# ---------------------------- Sheet-Zähler für dieses Blatt ------------------------------- #
            sheet = {key: 0 for key in total.keys()}


            for row in sheet_df.itertuples():

                # 1) PLZ + Stadt fehlen komplett
                if row.Postleitzahl == "NULLDATA" and row.Stadt == "NULLDATA":
                    sheet["NULLDATA"] += 1

                # 2) Polen
                elif row.Land == "PL":
                    sheet["polen"] += 1

                # 3) Deutschland → Unterkategorien
                elif row.Land == "DE":
                    if row.federalState == "Berlin":
                        sheet["berlin"] += 1
                    elif row.district == "Frankfurt (Oder)":
                        sheet["ffo"] += 1
                    elif row.municipality == "Schöneiche bei Berlin":
                        sheet["schoeneiche"] += 1
                    elif row.municipality == "Erkner":
                        sheet["erkner"] += 1
                    elif row.municipality == "Woltersdorf":
                        sheet["woltersdorf"] += 1
                    elif row.municipality == "Grünheide (Mark)":
                        sheet["gruenheide"] += 1
                    elif row.municipality == "Fürstenwalde/Spree":
                        sheet["fuewa"] += 1
                    elif row.municipality == "Spreenhagen":
                        sheet["spreenhagen"] += 1
                    elif row.municipality == "Bad Saarow":
                        sheet["bad_saarow"] += 1
                    elif row.municipality == "Storkow (Mark)":
                        sheet["storkow"] += 1
                    elif row.municipality == "Rietz-Neuendorf":
                        sheet["rietz_neuendorf"] += 1
                    elif row.municipality == "Tauche":
                        sheet["tauche"] += 1
                    elif row.municipality in [
                        "Berkenbrück",
                        "Briesen (Mark)",
                        "Jacobsdorf",
                        "Steinhöfel",
                    ]:
                        sheet["odervorland"] += 1
                    else:
                        sheet["sonstiges Deutschland"] += 1
                # 4) Foreign → nur wenn Land nicht NULLDATA
                elif row.Land != "NULLDATA":
                    sheet["foreign"] += 1

                # 5) Falls etwas NICHT erfasst wurde
                else:
                    sheet["no_category"] += 1


 # ---------------------------- Sheet-Ergebnis ausgeben ------------------------------- #

            sheet_total = sum(sheet.values())
            row_count = len(sheet_df)


            inner_circle = sum([sheet["schoeneiche"], sheet["erkner"], sheet["woltersdorf"], sheet["gruenheide"], sheet["fuewa"],
                                sheet["spreenhagen"], sheet["bad_saarow"], sheet["storkow"]])
            outer_circle = sum([sheet["rietz_neuendorf"], sheet["tauche"], sheet["odervorland"]])

            result_row = {
                "year": sheet_name,
                **sheet,
                "sum": sheet_total,
                "rows": row_count,
                "enger Bezugsraum": inner_circle,
                "weiterer Bezugsraum": outer_circle
            }
            results.append(result_row)

        self.analysis_df = pd.DataFrame(results)

        wb = load_workbook(self.excel_file_input)
        if "analysis" in wb.sheetnames:
            ws = wb["analysis"]
            wb.remove(ws)
        ws = wb.create_sheet("analysis")

        for r in dataframe_to_rows(self.analysis_df, index=False, header=True):
            ws.append(r)

        wb.save(self.excel_file_input)


        messagebox.showinfo(message="All sheets converted")


